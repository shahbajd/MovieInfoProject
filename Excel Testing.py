#!/usr/bin/env pythonimport osfrom openpyxl import load_workbook#Initialize Categoriestitle_col = 2year_col = 3rating_col = 4director_col = 5row_count = 3title = 'Testing 1'director = 'Director'year = 12345os.chdir('/Users/Shahbaj/Komodo10/Movie Info Retreival/')print os.getcwd()wb = load_workbook('Testing.xlsx')sheet_list = wb.get_sheet_names()active_sheet=wb.get_sheet_by_name(sheet_list[0])while(active_sheet.cell(row=row_count, column=title_col).value != None):	print active_sheet.cell(row=row_count, column=title_col).value	row_count+=1active_sheet.cell(row=row_count, column=title_col).value = titleactive_sheet.cell(row=row_count, column=year_col).value = yearactive_sheet.cell(row=row_count, column=director_col).value = directorwb.save('Testing.xlsx')#Create a duplicate checking system